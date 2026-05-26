# app/api/validation_routes.py

import io
import logging
import os
import tempfile
import threading
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
import pandas as pd
from fastapi import APIRouter, BackgroundTasks, File, Form, HTTPException, Query, UploadFile
from fastapi.responses import FileResponse

from app.core.config import settings
from app.schemas.validation_schema import (
    AsyncJobStartResponse,
    ColumnMatchItem,
    ColumnPreviewResponse,
    FacilitySummaryResponse,
    HealthResponse,
    JobStatusResponse,
    ReportMetaResponse,
    RuleDefinition,
    RuleListResponse,
    ValidationResponse,
    ViolationsPageResponse,
    build_validation_response,
    violations_df_to_records,
)
from app.services.validation_service import ValidationService
from app.utils.excel_helpers import ALL_REQUIRED_COLUMNS, CANONICAL_COLUMNS, build_column_rename_map

# ── Central-service additions ─────────────────────────────────────────────────
# Imported lazily-safe: if DB isn't initialised yet these still import fine.
from app.db.repository import RunRepository, ErrorRepository
from app.scheduler.setup import get_next_run_time

logger = logging.getLogger(__name__)

router = APIRouter()

_job_store: dict[str, dict] = {}


# ── Helpers ────────────────────────────────────────────────────────────────────

def _update_job(job_id: str, **kwargs) -> None:
    if job_id in _job_store:
        _job_store[job_id].update(kwargs)


async def _stream_to_disk(file: UploadFile, max_mb: int) -> Path:
    max_bytes = max_mb * 1024 * 1024
    total = 0
    suffix = Path(file.filename or "upload.xlsx").suffix or ".xlsx"
    tmp_fd, tmp_path_str = tempfile.mkstemp(suffix=suffix)
    tmp_path = Path(tmp_path_str)
    try:
        with os.fdopen(tmp_fd, "wb") as f:
            while True:
                chunk = await file.read(1024 * 1024)
                if not chunk:
                    break
                total += len(chunk)
                if total > max_bytes:
                    raise HTTPException(
                        status_code=413,
                        detail=(
                            f"File too large ({total / 1_048_576:.1f} MB). "
                            f"Maximum allowed: {max_mb} MB."
                        ),
                    )
                f.write(chunk)
    except HTTPException:
        tmp_path.unlink(missing_ok=True)
        raise
    except Exception as exc:
        tmp_path.unlink(missing_ok=True)
        raise HTTPException(status_code=500, detail=f"Upload stream failed: {exc}")
    return tmp_path


def _run_validation_task(
    job_id: str,
    tmp_path: Path,
    filename: str,
    rule_ids: list[str] | None,
    rep_start: pd.Timestamp | None,
    rep_end: pd.Timestamp | None,
) -> None:
    """
    Runs in anyio thread pool via BackgroundTask.

    CHANGE: generates the report to disk immediately after validation
    completes (eager generation). This means the .xlsx file exists on
    disk before Flutter opens the report screen, and the report endpoint
    can serve it without _job_store — surviving any backend restart.
    """
    stop_sim = threading.Event()

    def _simulate_progress() -> None:
        pct = 20
        while not stop_sim.wait(timeout=4.0):
            pct = min(pct + 4, 88)
            _job_store[job_id]["progress"] = pct
            if pct >= 88:
                break

    try:
        _update_job(job_id, status="processing", progress=5, message="Loading file…")

        file_bytes = tmp_path.read_bytes()
        size_mb = len(file_bytes) / 1_048_576
        enabled_count = len(rule_ids) if rule_ids else len(ValidationService.get_rules())

        _update_job(
            job_id,
            progress=15,
            message=f"Loaded {size_mb:.1f} MB — running {enabled_count} rules…",
        )

        sim = threading.Thread(target=_simulate_progress, daemon=True)
        sim.start()

        result = ValidationService.run(
            file_bytes=file_bytes,
            filename=filename,
            enabled_rule_ids=rule_ids,
            reporting_start=rep_start,
            reporting_end=rep_end,
        )

        stop_sim.set()

        # ── CHANGE: Generate report to disk immediately ────────────────────
        # The file is written before Flutter opens the report screen.
        # Even if the backend restarts after this, the file persists on disk
        # and GET /report/{job_id} can serve it without _job_store.
        _update_job(job_id, progress=92, message="Generating Excel report…")
        try:
            report_path = _get_report_path(job_id)
            if not report_path.exists():
                _generate_report(job_id, result)
            logger.info(f"Job {job_id}: report written → {report_path}")
        except Exception as report_err:
            # Non-fatal — validation succeeded; report will be tried on demand
            logger.warning(f"Job {job_id}: eager report generation failed: {report_err}")

        _update_job(
            job_id,
            status="done",
            progress=100,
            message=(
                f"Complete — {result.total_violations:,} violations "
                f"across {result.rules_run} rules "
                f"in {result.validation_time_seconds:.1f}s"
            ),
            result=result,
        )

        # Evict old done/error desktop-upload jobs to prevent DataFrame accumulation.
        # Keep at most 20 entries; the report file on disk is the durable artifact.
        _evict_old_desktop_jobs(keep=20, exclude={job_id})

        logger.info(
            f"Job {job_id}: done — "
            f"{result.total_violations} violations, "
            f"{result.total_rows} rows, "
            f"{result.validation_time_seconds:.1f}s"
        )

    except Exception as exc:
        stop_sim.set()
        logger.exception(f"Job {job_id}: validation failed — {exc}")
        _update_job(job_id, status="error", progress=0, message=str(exc))

    finally:
        stop_sim.set()
        try:
            tmp_path.unlink(missing_ok=True)
        except Exception:
            pass


# ═══════════════════════════════════════════════════════════════════════════════
# GET /health
# ═══════════════════════════════════════════════════════════════════════════════

@router.get("/health", response_model=HealthResponse, tags=["System"])
def health_check() -> dict:
    """BayCentral calls this on every page navigation.
    Returns last completed run info and next scheduled run time.
    A 200 means the service is online; BayCentral shows the EMR Validation menu.
    """
    last_run_info = None
    try:
        last = RunRepository.get_last_completed()
        if last:
            last_run_info = {
                "run_id":       last.run_id,
                "schedule_slot": last.schedule_slot,
                "run_time":     last.run_time.strftime("%Y-%m-%dT%H:%M:%SZ") if last.run_time else None,
                "status":       last.status,
            }
    except Exception:
        pass  # DB may not be ready on very first startup — still return 200

    return {
        "status":       "ok",
        "version":      settings.app_version,
        "last_run":     last_run_info,
        "next_run_at":  get_next_run_time(),
    }


# ═══════════════════════════════════════════════════════════════════════════════
# GET /runs  — BayCentral home page data source
# Returns the latest completed run per schedule slot (at most 3 rows).
# When a new run for a slot completes it overwrites the previous one here.
# ═══════════════════════════════════════════════════════════════════════════════

@router.get("/runs", tags=["Runs"])
def list_runs() -> list:
    """Latest completed run per slot — what BayCentral shows on its home page."""
    runs = RunRepository.get_latest_per_slot()
    return [
        {
            "run_id":                  r.run_id,
            "schedule_slot":           r.schedule_slot,
            "run_time":                r.run_time.strftime("%Y-%m-%dT%H:%M:%SZ") if r.run_time else None,
            "blob_date":               r.blob_date.isoformat() if r.blob_date else None,
            "status":                  r.status,
            "total_rows":              r.total_rows,
            "total_violations":        r.total_violations,
            "rules_run":               r.rules_run,
            "validation_time_seconds": r.validation_time_seconds,
            **r.get_result_json(),    # inlines rule_summaries + facility_summaries
        }
        for r in runs
    ]


# ═══════════════════════════════════════════════════════════════════════════════
# GET /runs/history  — lightweight history for the trend chart
# Returns all completed runs (no rule/facility summaries) newest-first.
# ═══════════════════════════════════════════════════════════════════════════════

@router.get("/runs/history", tags=["Runs"])
def list_runs_history(days: int = Query(default=90, ge=1, le=365)) -> list:
    """All completed runs within the last `days` days.
    Intentionally excludes rule_summaries / facility_summaries to keep
    the payload small — the trend chart only needs run_time + total_violations.
    Returns an empty list on DB error so the frontend degrades gracefully.
    """
    try:
        runs = RunRepository.get_all_completed(days=days)
        return [
            {
                "run_id":           r.run_id,
                "schedule_slot":    r.schedule_slot,
                "run_time":         r.run_time.strftime("%Y-%m-%dT%H:%M:%SZ") if r.run_time else None,
                "total_violations": r.total_violations,
                "total_rows":       r.total_rows,
                "status":           r.status,
            }
            for r in runs
        ]
    except Exception as exc:
        logger.warning(f"GET /runs/history failed — returning empty list: {exc}")
        return []


# ═══════════════════════════════════════════════════════════════════════════════
# GET /rules
# ═══════════════════════════════════════════════════════════════════════════════

@router.get("/rules", response_model=RuleListResponse, tags=["Rules"])
def list_rules() -> RuleListResponse:
    try:
        rules_data = ValidationService.get_rules()
    except FileNotFoundError as e:
        raise HTTPException(status_code=500, detail=str(e))

    rule_defs = [
        RuleDefinition(
            id=r["id"],
            title=r["title"],
            description=r.get("description", ""),
            category=r.get("category", ""),
            severity=r["severity"],
            enabled=r.get("enabled", True),
            columns=r.get("columns", []),
        )
        for r in rules_data
    ]
    return RuleListResponse(total=len(rule_defs), rules=rule_defs)


# ═══════════════════════════════════════════════════════════════════════════════
# POST /columns/preview
# ═══════════════════════════════════════════════════════════════════════════════

@router.post("/columns/preview", response_model=ColumnPreviewResponse, tags=["Validation"])
async def preview_columns(file: UploadFile = File(...)) -> ColumnPreviewResponse:

    filename = file.filename or ""
    if not filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail=f"Only .xlsx files accepted. Got: '{filename}'")

    file_bytes = await file.read()
    if not file_bytes:
        raise HTTPException(status_code=400, detail="Uploaded file is empty.")

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"Could not open Excel workbook: {e}")

    sheet_name = wb.sheetnames[0] if wb.sheetnames else "Sheet1"
    ws = wb[sheet_name]
    header_row: list[str] = []
    try:
        for row in ws.iter_rows(max_row=10, values_only=True):
            cells = [str(c).strip() for c in row if c is not None and str(c).strip()]
            if len(cells) >= 3:
                header_row = cells
                break
    finally:
        wb.close()

    if not header_row:
        raise HTTPException(
            status_code=422,
            detail=f"No header row found in sheet '{sheet_name}'.",
        )

    rename_map = build_column_rename_map(header_row, threshold=80)
    exact_lookup: set[str] = set()
    for variations in CANONICAL_COLUMNS.values():
        for v in variations:
            exact_lookup.add(v.lower().strip())

    claimed_canonicals = set(rename_map.values())
    matched_items = [
        ColumnMatchItem(
            file_header=fh,
            canonical=canon,
            score=100.0 if fh.lower().strip() in exact_lookup else 80.0,
        )
        for fh, canon in rename_map.items()
    ]
    missing_canonicals = [c for c in ALL_REQUIRED_COLUMNS if c not in claimed_canonicals]
    extra_file_headers = [h for h in header_row if h not in rename_map]
    coverage_pct = round(
        len(claimed_canonicals) / len(ALL_REQUIRED_COLUMNS) * 100, 1
    ) if ALL_REQUIRED_COLUMNS else 100.0

    return ColumnPreviewResponse(
        sheet_name=sheet_name,
        total_in_file=len(header_row),
        total_required=len(ALL_REQUIRED_COLUMNS),
        coverage_pct=coverage_pct,
        matched=matched_items,
        missing_canonicals=missing_canonicals,
        extra_file_headers=extra_file_headers,
    )


# ═══════════════════════════════════════════════════════════════════════════════
# POST /validate
# ═══════════════════════════════════════════════════════════════════════════════

@router.post(
    "/validate",
    response_model=AsyncJobStartResponse,
    tags=["Validation"],
)
async def validate_radet(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    enabled_rules: str | None = Form(default=None),
    reporting_start: str | None = Form(default=None),
    reporting_end: str | None = Form(default=None),
    page_size: int = Form(default=200, ge=1, le=1000),
) -> AsyncJobStartResponse:

    filename = file.filename or ""
    if not filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(
            status_code=400,
            detail=f"Invalid file type: '{filename}'. Only .xlsx files accepted.",
        )

    tmp_path = await _stream_to_disk(file, max_mb=settings.max_upload_mb)
    logger.info(
        f"Upload received: '{filename}' → {tmp_path} "
        f"({tmp_path.stat().st_size / 1_048_576:.2f} MB)"
    )

    rule_ids: list[str] | None = None
    if enabled_rules:
        rule_ids = [r.strip() for r in enabled_rules.split(",") if r.strip()]

    rep_start: pd.Timestamp | None = None
    rep_end: pd.Timestamp | None = None

    if reporting_start:
        try:
            rep_start = pd.Timestamp(reporting_start)
        except Exception:
            raise HTTPException(400, detail=f"Invalid reporting_start: '{reporting_start}'.")

    if reporting_end:
        try:
            rep_end = pd.Timestamp(reporting_end)
        except Exception:
            raise HTTPException(400, detail=f"Invalid reporting_end: '{reporting_end}'.")

    if rep_start and rep_end and rep_end < rep_start:
        raise HTTPException(400, detail="reporting_end must be on or after reporting_start.")

    from uuid import uuid4
    job_id = str(uuid4())[:8]

    _job_store[job_id] = {
        "status": "pending",
        "progress": 0,
        "message": "Queued — waiting for engine",
        "result": None,
    }

    background_tasks.add_task(
        _run_validation_task,
        job_id, tmp_path, filename, rule_ids, rep_start, rep_end,
    )

    logger.info(f"Job {job_id} queued for '{filename}'")
    return AsyncJobStartResponse(job_id=job_id, status="pending")


# ═══════════════════════════════════════════════════════════════════════════════
# GET /validate/{job_id}/status
# ═══════════════════════════════════════════════════════════════════════════════

@router.get("/validate/{job_id}/status", response_model=JobStatusResponse, tags=["Validation"])
def get_job_status(job_id: str) -> JobStatusResponse:
    job = _job_store.get(job_id)
    if job is None:
        raise HTTPException(
            status_code=404,
            detail=f"Job '{job_id}' not found. It may have expired or never existed.",
        )
 
    status     = job.get("status", "unknown")
    progress   = job.get("progress", 0)
    message    = job.get("message", "")
    result_obj = job.get("result")
 
    response_result: ValidationResponse | None = None
    if status == "done" and result_obj is not None:
        # ── FIX: pass the API job_id so ValidationResponse.job_id == the
        #         report filename and _job_store key.
        #         Without this, result_obj.job_id is the *internal* validator
        #         job_id which differs from the API job_id, causing Flutter
        #         to request GET /report/<wrong_id> → 404.
        response_result = build_validation_response(
            result_obj,
            page=1,
            page_size=200,
            override_job_id=job_id,   # ← THE FIX
        )
 
    return JobStatusResponse(
        job_id=job_id,
        status=status,
        progress=progress,
        message=message,
        result=response_result,
    )

# ═══════════════════════════════════════════════════════════════════════════════
# GET /validate/{job_id}/violations
# ═══════════════════════════════════════════════════════════════════════════════

@router.get("/validate/{job_id}/violations", response_model=ViolationsPageResponse, tags=["Validation"])
def get_violations_page(
    job_id: str,
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=200, ge=1, le=1000),
    rule_id: str | None = Query(default=None),
    facility: str | None = Query(default=None),
    lga: str | None = Query(default=None),
    state: str | None = Query(default=None),
    facilities: str | None = Query(default=None),  # comma-separated list for state filter
    severity: str | None = Query(default=None),
) -> ViolationsPageResponse:
    """Paginated violations.
    Checks _job_store first (fast, in-memory) then falls back to SQLite
    (survives service restart).  BayCentral uses this endpoint.
    """

    # ── Path A: in-memory job store (desktop runs + same-session central runs) ─
    job = _job_store.get(job_id)
    if job is not None:
        if job.get("status") != "done":
            raise HTTPException(
                status_code=400,
                detail=f"Job '{job_id}' is not complete yet (status: {job.get('status')}).",
            )
        result = job.get("result")
        if result is not None:
            violations_df = result.violations_df.copy()
            if rule_id:
                violations_df = violations_df[
                    violations_df["rule_id"].str.upper() == rule_id.upper()
                ]
            if facility:
                violations_df = violations_df[
                    violations_df["facility_name"].str.lower().str.contains(facility.lower(), na=False)
                ]
            if lga:
                violations_df = violations_df[
                    violations_df["lga_of_residence"].str.lower().str.contains(lga.lower(), na=False)
                ]
            if state and "state" in violations_df.columns:
                violations_df = violations_df[
                    violations_df["state"].str.lower() == state.lower()
                ]
            if facilities:
                fac_list = [f.strip() for f in facilities.split(",") if f.strip()]
                if fac_list:
                    violations_df = violations_df[
                        violations_df["facility_name"].isin(fac_list)
                    ]
            if severity:
                violations_df = violations_df[
                    violations_df["severity"].str.lower() == severity.lower()
                ]
            records, total_pages, total_records = violations_df_to_records(
                violations_df, page=page, page_size=page_size
            )
            return ViolationsPageResponse(
                job_id=job_id,
                page=page,
                page_size=page_size,
                total_pages=total_pages,
                total_records=total_records,
                violations=records,
            )

    # ── Path B: SQLite fallback (after service restart, or BayCentral reading ──
    #            older scheduled runs not in memory anymore)
    try:
        fac_list = [f.strip() for f in facilities.split(",") if f.strip()] if facilities else None
        db_result = ErrorRepository.paginate(
            run_id=job_id,
            page=page,
            page_size=page_size,
            rule_id=rule_id,
            facility=facility,
            lga=lga,
            state=state,
            facilities=fac_list,
            severity=severity,
        )
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Database error: {exc}")

    if db_result["total_records"] == 0:
        # Check whether the run exists at all vs just zero violations
        run = RunRepository.get_by_run_id(job_id)
        if run is None:
            raise HTTPException(status_code=404, detail=f"Job '{job_id}' not found.")

    return ViolationsPageResponse(
        job_id=job_id,
        page=page,
        page_size=page_size,
        total_pages=db_result["total_pages"],
        total_records=db_result["total_records"],
        violations=db_result["violations"],
    )



# ═══════════════════════════════════════════════════════════════════════════════
# GET /report/{job_id}/meta
#
# CHANGE: checks disk FIRST — no _job_store needed if file already exists.
# The report is generated eagerly in _run_validation_task(), so the file
# should be on disk before Flutter opens the report screen.
# This survives backend restarts because the file persists in reports/.
# ═══════════════════════════════════════════════════════════════════════════════

@router.get("/report/{job_id}/meta", response_model=ReportMetaResponse, tags=["Report"])
def get_report_meta(job_id: str) -> ReportMetaResponse:
    report_path = _get_report_path(job_id)

    # ── Path A: file already on disk (normal case after eager generation) ─────
    if report_path.exists():
        stat = report_path.stat()
        return ReportMetaResponse(
            job_id=job_id,
            file_name=report_path.name,
            file_size_kb=round(stat.st_size / 1024, 1),
            generated_at=datetime.fromtimestamp(stat.st_mtime, tz=timezone.utc),
            total_sheets=4,
        )

    # ── Path B: file not on disk — try to generate from _job_store ───────────
    job = _job_store.get(job_id)
    if job is None:
        raise HTTPException(
            status_code=404,
            detail=(
                f"Report not found for job '{job_id}'. "
                "The engine may have restarted and lost the job. "
                "Please go back to Home and re-run validation."
            ),
        )

    result = job.get("result")
    if result is None:
        raise HTTPException(
            status_code=400,
            detail=f"Job '{job_id}' is not complete yet (status: {job.get('status')}).",
        )

    _generate_report(job_id, result)
    stat = report_path.stat()
    return ReportMetaResponse(
        job_id=job_id,
        file_name=report_path.name,
        file_size_kb=round(stat.st_size / 1024, 1),
        generated_at=datetime.fromtimestamp(stat.st_mtime, tz=timezone.utc),
        total_sheets=4,
    )


# ═══════════════════════════════════════════════════════════════════════════════
# GET /report/{job_id}
#
# CHANGE: checks disk FIRST — same pattern as get_report_meta().
# ═══════════════════════════════════════════════════════════════════════════════

@router.get("/report/{job_id}", tags=["Report"])
def download_report(job_id: str):
    report_path = _get_report_path(job_id)

    # ── Path A: serve directly from disk ──────────────────────────────────────
    if report_path.exists():
        return FileResponse(
            path=str(report_path),
            filename=report_path.name,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    # ── Path B: file missing — try _job_store ─────────────────────────────────
    job = _job_store.get(job_id)
    if job is not None:
        result = job.get("result")
        if result is None:
            raise HTTPException(
                status_code=400,
                detail=f"Job '{job_id}' not ready (status: {job.get('status')}).",
            )
        try:
            _generate_report(job_id, result)
        except Exception as e:
            logger.exception(f"Report generation failed for job {job_id}: {e}")
            raise HTTPException(status_code=500, detail=f"Report generation failed: {e}")
        return FileResponse(
            path=str(report_path),
            filename=report_path.name,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    # ── Path C: not in memory — try rebuilding from SQLite ───────────────────
    result = _rebuild_result_from_db(job_id)
    if result is None:
        raise HTTPException(
            status_code=404,
            detail=f"Report not found for job '{job_id}'. Run not in database.",
        )
    try:
        _generate_report(job_id, result)
    except Exception as e:
        logger.exception(f"Report regeneration from DB failed for job {job_id}: {e}")
        raise HTTPException(status_code=500, detail=f"Report generation failed: {e}")

    return FileResponse(
        path=str(report_path),
        filename=report_path.name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ─── Private helpers ──────────────────────────────────────────────────────────

def _evict_old_desktop_jobs(keep: int, exclude: set) -> None:
    """Remove oldest done/error entries from _job_store, keeping at most `keep`."""
    evictable = [
        k for k, v in _job_store.items()
        if k not in exclude and v.get("status") in ("done", "error")
    ]
    excess = len(_job_store) - keep
    if excess > 0:
        for k in evictable[:excess]:
            del _job_store[k]


def _get_report_path(job_id: str) -> Path:
    report_dir = Path(settings.report_dir)
    report_dir.mkdir(parents=True, exist_ok=True)
    return report_dir / f"EMR_Validation_Report_{job_id}.xlsx"


def _generate_report(job_id: str, validation_result) -> None:
    from app.engine.reporter import generate_report
    report_path = _get_report_path(job_id)
    generate_report(validation_result, output_path=report_path)
    logger.info(f"Report generated: {report_path}")


def _rebuild_result_from_db(job_id: str):
    """Reconstruct a minimal result object from SQLite so the report can be
    regenerated after a service restart wiped _job_store."""
    from types import SimpleNamespace
    from app.db.models import FacilityError
    from app.db.session import get_db

    run = RunRepository.get_by_run_id(job_id)
    if run is None or run.status != "done":
        return None

    # Rebuild violations DataFrame from facility_errors
    with get_db() as db:
        errors = (
            db.query(FacilityError)
            .filter(FacilityError.run_id_fk == run.id)
            .all()
        )
        rows = [
            {
                "_row_number":        e.row_number or 0,
                "patient_id":         e.patient_id or "",
                "patient_uid":        e.patient_uid or "",
                "facility_name":      e.facility_name or "",
                "state":              e.state or "",
                "lga_of_residence":   e.lga_of_residence or "",
                "rule_id":            e.rule_id or "",
                "rule_title":         e.rule_title or "",
                "severity":           e.severity or "",
                "failing_field":      e.failing_field or "",
                "current_value":      e.current_value or "",
                "expected_condition": e.expected_condition or "",
            }
            for e in errors
        ]

    violations_df = pd.DataFrame(rows) if rows else pd.DataFrame(columns=[
        "_row_number", "patient_id", "patient_uid", "facility_name", "state",
        "lga_of_residence", "rule_id", "rule_title", "severity",
        "failing_field", "current_value", "expected_condition",
    ])

    result_data = run.get_result_json()
    rule_sums = [SimpleNamespace(**rs) for rs in result_data.get("rule_summaries", [])]
    fac_sums  = [SimpleNamespace(**fs) for fs in result_data.get("facility_summaries", [])]
    rules_skipped = sum(1 for rs in rule_sums if getattr(rs, "skipped", False))

    logger.info(
        f"Rebuilding report for run {job_id} from DB: "
        f"{run.total_violations} violations, {len(rule_sums)} rules"
    )

    return SimpleNamespace(
        job_id=job_id,
        total_rows=run.total_rows or 0,
        total_violations=run.total_violations or 0,
        rules_run=run.rules_run or 0,
        rules_skipped=rules_skipped,
        validation_time_seconds=run.validation_time_seconds or 0.0,
        engine_warnings=[],
        rule_summaries=rule_sums,
        facility_summaries=fac_sums,
        violations_df=violations_df,
    )
